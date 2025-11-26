from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('K-8 Worksheet Topics List.xlsx')
ws = wb['Grade 5']

print("Grade 5 Topics:")
print("="*80)

current_unit = None
topics_by_unit = {}

for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
    if row[0] is None:
        continue

    unit = int(row[0])
    topic_type = row[1]
    topic = row[2]

    if unit not in topics_by_unit:
        topics_by_unit[unit] = []

    topics_by_unit[unit].append({
        'type': topic_type,
        'topic': topic
    })

# Print organized by unit
for unit in sorted(topics_by_unit.keys()):
    print(f"\nUnit {unit}:")
    for topic_info in topics_by_unit[unit]:
        print(f"  - {topic_info['topic']} ({topic_info['type']})")
