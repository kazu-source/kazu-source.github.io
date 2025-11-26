from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('K-8 Worksheet Topics List.xlsx')

# Check if there's a Grade 5 or Grade_5 sheet
print("Available sheets:", wb.sheetnames)
print("\n" + "="*50)

# Try to find Grade 5 data
for sheet_name in wb.sheetnames:
    if '5' in sheet_name or 'five' in sheet_name.lower():
        print(f"\nFound sheet: {sheet_name}")
        ws = wb[sheet_name]
        print("\nTopics:")
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            if any(cell for cell in row):  # Skip empty rows
                print(row)
