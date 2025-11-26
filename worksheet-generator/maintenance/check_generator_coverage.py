"""
Generator Coverage Checker

This script checks if all topics listed in an Excel file have corresponding
generator files in the generators directory.

Usage:
    python check_generator_coverage.py <excel_file_path> [--output-report]

Examples:
    python check_generator_coverage.py "../High School Worksheet Topics List.xlsx"
    python check_generator_coverage.py "../Middle School Worksheet Topics List.xlsx" --output-report
    python check_generator_coverage.py "../Elementary Worksheet Topics List.xlsx" --output-report
"""

import os
import sys
import argparse
from openpyxl import load_workbook
import re
from datetime import datetime
import json


def load_config(config_path):
    """Load configuration from JSON file."""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)

        # Set defaults for any missing values
        defaults = {
            'grade_level': 'Unknown',
            'folder_pattern': 'chapter',
            'excluded_types': ['Expansion', 'Review'],
            'unit_format': 'numeric',
            'generators_subdir': 'generators'
        }

        for key, default_value in defaults.items():
            if key not in config:
                config[key] = default_value

        return config
    except FileNotFoundError:
        raise FileNotFoundError(f"Config file not found: {config_path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in config file: {e}")


def topic_to_generator_name(topic):
    """Convert topic name to expected generator filename."""
    # Convert topic to lowercase, replace spaces and special chars with underscores
    name = topic.lower()
    # Remove parentheses content but keep the text
    name = re.sub(r'\s*\([^)]*\)', '', name)
    # Replace various punctuation and spaces
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[-\s]+', '_', name)
    name = name.strip('_')
    return name + '_generator.py'


def get_generator_files(generators_dir, folder_pattern='chapter'):
    """Scan the generators directory and return all generator files by unit folder.

    Args:
        generators_dir: Path to the generators directory
        folder_pattern: Prefix pattern for unit folders (e.g., 'chapter', 'lesson', 'unit')
    """
    generator_files = {}

    if not os.path.exists(generators_dir):
        print(f"Warning: Generators directory not found: {generators_dir}")
        return generator_files

    for folder in os.listdir(generators_dir):
        folder_path = os.path.join(generators_dir, folder)
        if os.path.isdir(folder_path) and folder.startswith(folder_pattern):
            # Extract the unit number from folder name (e.g., 'chapter01' -> '01', 'lesson5' -> '5')
            unit_num = folder.replace(folder_pattern, '')
            generator_files[unit_num] = []
            for file in os.listdir(folder_path):
                if file.endswith('.py') and file != '__init__.py':
                    generator_files[unit_num].append(file)

    return generator_files


def load_topics_from_excel(excel_path, sheet_name=None):
    """Load topics from the Excel file.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Optional sheet name to read from (if None, uses active sheet)
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    topics = []
    current_unit = None

    for row in ws.iter_rows(min_row=2):
        unit = row[0].value
        type_val = row[1].value
        topic = row[2].value

        # Track current unit (for cases where unit cell is merged or empty)
        if unit and str(unit).lower() not in ['unit', 'type']:
            try:
                current_unit = str(int(float(unit)))
            except:
                current_unit = str(unit)

        # Add topic if we have both type and topic (and skip headers)
        if type_val and topic and str(type_val).lower() not in ['type', 'unit']:
            topics.append({
                'unit': current_unit if current_unit else '',
                'type': type_val.strip() if type_val else '',
                'topic': topic.strip() if topic else ''
            })

    return topics


def fuzzy_match_generator(expected_name, actual_files):
    """
    Improved fuzzy matching for generator files.
    Returns the best matching file or None.
    """
    expected_base = expected_name.replace('_generator.py', '')
    expected_words = set(expected_base.split('_'))

    # Common filler words to ignore in matching
    filler_words = {'of', 'the', 'a', 'an', 'and', 'or', 'in', 'on', 'at', 'to', 'for', 'with'}
    expected_words_filtered = expected_words - filler_words

    best_match = None
    best_score = 0

    for actual_file in actual_files:
        actual_base = actual_file.replace('_generator.py', '').replace('.py', '')
        actual_words = set(actual_base.split('_'))
        actual_words_filtered = actual_words - filler_words

        # Calculate similarity score
        # 1. Check for exact match first
        if expected_base == actual_base:
            return actual_file

        # 2. Check if expected is substring of actual OR actual is substring of expected
        if expected_base in actual_base or actual_base in expected_base:
            score = min(len(expected_base), len(actual_base)) / max(len(expected_base), len(actual_base))
            if score > best_score and score > 0.4:  # Lowered threshold for substring matches
                best_score = score
                best_match = actual_file
            continue

        # 3. Calculate word overlap score (using filtered words)
        common_words = expected_words_filtered & actual_words_filtered
        if common_words and expected_words_filtered:
            # Score based on percentage of matching significant words
            score = len(common_words) / max(len(expected_words_filtered), len(actual_words_filtered))

            # Bonus for matching key words
            key_words = {'solving', 'graphing', 'equations', 'inequalities', 'systems', 'quadratic', 'linear',
                         'functions', 'exponential', 'polynomial', 'factoring', 'vertex', 'slope', 'intercept'}
            key_matches = common_words & key_words
            if key_matches:
                score += 0.2 * len(key_matches)

            if score > best_score and score >= 0.5:  # Require at least 50% match
                best_score = score
                best_match = actual_file

        # 4. Special case: if all actual words are in expected words (ignoring fillers), it's likely a match
        if actual_words_filtered and actual_words_filtered.issubset(expected_words_filtered):
            score = 0.8  # High score for subset matches
            if score > best_score:
                best_score = score
                best_match = actual_file

    return best_match if best_score > 0.4 else None  # Lowered overall threshold


def check_coverage(topics, generator_files, excluded_types=None):
    """Check which topics have generators and which are missing.

    Args:
        topics: List of topic dictionaries from Excel
        generator_files: Dictionary of generator files by unit
        excluded_types: List of topic types to exclude from coverage calculation (e.g., ['Expansion', 'Review'])
    """
    if excluded_types is None:
        excluded_types = ['Expansion', 'Review']

    missing = []
    found = []
    excluded_topics = {type_name: [] for type_name in excluded_types}

    for topic_info in topics:
        unit = topic_info['unit']
        type_val = topic_info['type']
        topic = topic_info['topic']

        # Skip excluded topic types from coverage calculation
        if type_val in excluded_types:
            excluded_topics[type_val].append(topic_info)
            continue

        expected_filename = topic_to_generator_name(topic)

        # Check if generator exists in the corresponding chapter
        chapter_key = unit.zfill(2) if unit.isdigit() else unit

        if chapter_key in generator_files:
            if expected_filename in generator_files[chapter_key]:
                found.append({**topic_info, 'filename': expected_filename})
            else:
                # Try to find a similar filename using improved fuzzy matching
                matched_file = fuzzy_match_generator(expected_filename, generator_files[chapter_key])
                if matched_file:
                    found.append({**topic_info, 'filename': matched_file})
                else:
                    missing.append({**topic_info, 'expected': expected_filename})
        else:
            missing.append({**topic_info, 'expected': expected_filename})

    return found, missing, excluded_topics


def print_report(topics, generator_files, found, missing, excluded_topics, excel_filename, config):
    """Print the coverage report to console.

    Args:
        topics: All topics from Excel
        generator_files: Dictionary of generator files by unit
        found: List of topics with generators found
        missing: List of topics with missing generators
        excluded_topics: Dictionary of excluded topics by type
        excel_filename: Name of the Excel file
        config: Configuration dictionary
    """
    folder_pattern = config.get('folder_pattern', 'chapter')
    grade_level = config.get('grade_level', 'Unknown')

    print(f"\n{'='*80}")
    print(f"GENERATOR COVERAGE REPORT - {grade_level}")
    print(f"Source: {excel_filename}")
    print(f"{'='*80}\n")

    print(f"Total topics in Excel: {len(topics)}\n")

    print(f"Generator files by {folder_pattern}:")
    for unit_num in sorted(generator_files.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        print(f"  {folder_pattern.capitalize()} {unit_num}: {len(generator_files[unit_num])} generators")
    print()

    excluded_count = sum(len(topics_list) for topics_list in excluded_topics.values())
    regular_topics = len(topics) - excluded_count

    print(f"\n{'='*80}")
    print(f"SUMMARY")
    print(f"{'='*80}")
    excluded_types_str = ' & '.join(config.get('excluded_types', []))
    print(f"Total topics (excluding {excluded_types_str}): {regular_topics}")
    print(f"Generators found: {len(found)}")
    print(f"Generators missing: {len(missing)}")
    for type_name, topics_list in excluded_topics.items():
        if topics_list:
            print(f"{type_name} topics (not checked): {len(topics_list)}")

    coverage_pct = (len(found) / regular_topics * 100) if regular_topics > 0 else 0
    print(f"Coverage: {coverage_pct:.1f}%")

    if missing:
        print(f"\n{'='*80}")
        print(f"MISSING GENERATORS ({len(missing)})")
        print(f"{'='*80}")
        for item in missing:
            unit_formatted = item['unit'].zfill(2) if item['unit'].isdigit() else item['unit']
            print(f"\nUnit {item['unit']} - {item['type']} - {item['topic']}")
            print(f"  Expected: {folder_pattern}{unit_formatted}/{item['expected']}")

    for type_name, topics_list in excluded_topics.items():
        if topics_list:
            print(f"\n{'='*80}")
            print(f"{type_name.upper()} TOPICS (not checked, {len(topics_list)})")
            print(f"{'='*80}")
            for item in topics_list:
                print(f"  Unit {item['unit']} - {item['topic']}")

    print(f"\n{'='*80}")
    print(f"FOUND GENERATORS ({len(found)})")
    print(f"{'='*80}")
    by_unit = {}
    for item in found:
        unit = item['unit']
        if unit not in by_unit:
            by_unit[unit] = []
        by_unit[unit].append(item)

    for unit in sorted(by_unit.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        print(f"\nUnit {unit}: {len(by_unit[unit])} generators")
        for item in by_unit[unit]:
            print(f"  [OK] {item['topic']} -> {item['filename']}")


def generate_markdown_report(topics, generator_files, found, missing, excluded_topics, excel_filename, output_path, config):
    """Generate a detailed markdown report.

    Args:
        topics: All topics from Excel
        generator_files: Dictionary of generator files by unit
        found: List of topics with generators found
        missing: List of topics with missing generators
        excluded_topics: Dictionary of excluded topics by type
        excel_filename: Name of the Excel file
        output_path: Path to write the markdown report
        config: Configuration dictionary
    """
    folder_pattern = config.get('folder_pattern', 'chapter')
    grade_level = config.get('grade_level', 'Unknown')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"# Generator Coverage Report - {grade_level}\n\n")
        f.write(f"**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**Source**: {excel_filename}\n\n")

        f.write(f"## Summary\n\n")
        excluded_count = sum(len(topics_list) for topics_list in excluded_topics.values())
        total_regular = len(topics) - excluded_count
        coverage_pct = (len(found) / total_regular * 100) if total_regular > 0 else 0

        f.write(f"- **Total topics**: {len(topics)}\n")
        f.write(f"- **Regular topics**: {total_regular}\n")
        for type_name, topics_list in excluded_topics.items():
            if topics_list:
                f.write(f"- **{type_name} topics**: {len(topics_list)}\n")
        f.write(f"- **Generators found**: {len(found)}\n")
        f.write(f"- **Generators missing**: {len(missing)}\n")
        f.write(f"- **Coverage**: {coverage_pct:.1f}%\n\n")

        f.write(f"## Generators by {folder_pattern.capitalize()}\n\n")
        for unit_num in sorted(generator_files.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            f.write(f"- {folder_pattern.capitalize()} {unit_num}: {len(generator_files[unit_num])} generators\n")

        if missing:
            f.write(f"\n## Missing Generators ({len(missing)})\n\n")
            f.write(f"| Unit | Type | Topic | Expected File |\n")
            f.write(f"|------|------|-------|---------------|\n")
            for item in missing:
                unit_formatted = item['unit'].zfill(2) if item['unit'].isdigit() else item['unit']
                f.write(f"| {item['unit']} | {item['type']} | {item['topic']} | {folder_pattern}{unit_formatted}/{item['expected']} |\n")

        for type_name, topics_list in excluded_topics.items():
            if topics_list:
                f.write(f"\n## {type_name} Topics ({len(topics_list)})\n\n")
                f.write(f"These are {type_name.lower()} topics that may not require worksheet generators.\n\n")
                f.write(f"| Unit | Topic |\n")
                f.write(f"|------|-------|\n")
                for item in topics_list:
                    f.write(f"| {item['unit']} | {item['topic']} |\n")

        f.write(f"\n## Found Generators by Unit\n\n")
        by_unit = {}
        for item in found:
            unit = item['unit']
            if unit not in by_unit:
                by_unit[unit] = []
            by_unit[unit].append(item)

        for unit in sorted(by_unit.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            f.write(f"\n### Unit {unit} ({len(by_unit[unit])} generators)\n\n")
            f.write(f"| Topic | Generator File |\n")
            f.write(f"|-------|----------------|\n")
            for item in by_unit[unit]:
                f.write(f"| {item['topic']} | {item['filename']} |\n")


def main():
    parser = argparse.ArgumentParser(
        description='Check generator coverage against Excel topic list',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python check_generator_coverage.py "../High School Worksheet Topics List.xlsx" --config config_highschool.json
  python check_generator_coverage.py "../K-8 Topics.xlsx" --config config_k8.json --output-report
        """
    )
    parser.add_argument('excel_file', help='Path to the Excel file containing topics')
    parser.add_argument('--config', '-c', required=True,
                       help='Path to configuration JSON file (e.g., config_highschool.json)')
    parser.add_argument('--output-report', '-o', action='store_true',
                       help='Generate a detailed markdown report file')
    parser.add_argument('--generators-dir', '-g', default=None,
                       help='Path to generators directory (overrides config file)')

    args = parser.parse_args()

    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Resolve paths relative to script directory
    excel_path = os.path.join(script_dir, args.excel_file)
    config_path = os.path.join(script_dir, args.config)

    # Get just the filename for display
    excel_filename = os.path.basename(args.excel_file)

    try:
        # Load configuration
        config = load_config(config_path)

        # Determine generators directory (command line overrides config)
        if args.generators_dir:
            generators_dir = os.path.join(script_dir, args.generators_dir)
        else:
            generators_subdir = config.get('generators_subdir', 'generators')
            generators_dir = os.path.join(script_dir, '..', generators_subdir)

        # Load topics from Excel (with optional sheet name from config)
        sheet_name = config.get('excel_sheet', None)
        topics = load_topics_from_excel(excel_path, sheet_name)

        # Get generator files using folder pattern from config
        folder_pattern = config.get('folder_pattern', 'chapter')
        generator_files = get_generator_files(generators_dir, folder_pattern)

        # Check coverage using excluded types from config
        excluded_types = config.get('excluded_types', ['Expansion', 'Review'])
        found, missing, excluded_topics = check_coverage(topics, generator_files, excluded_types)

        # Print console report
        print_report(topics, generator_files, found, missing, excluded_topics, excel_filename, config)

        # Generate markdown report if requested
        if args.output_report:
            # Create report filename based on Excel filename
            report_name = excel_filename.replace('.xlsx', '_coverage_report.md')
            report_path = os.path.join(script_dir, report_name)
            generate_markdown_report(topics, generator_files, found, missing,
                                   excluded_topics, excel_filename, report_path, config)
            print(f"\n{'='*80}")
            print(f"Markdown report saved to: {report_path}")
            print(f"{'='*80}\n")

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Configuration error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
